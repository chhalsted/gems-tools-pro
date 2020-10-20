#! python3

"""
Geologic Names Check
name-fullname - This version checks for Geolex names in both the name and fullname and then processes the set 
with the fewest entries to minimize the number of Geolex names and usages that are reported.

Arguments: 
    DMU - GeMS DescriptionOfMapUnits table. Geodatabase, CSV, tab delimeted TXT, or DBF. Required.
    Extent - one or more (comma separated) state or US region abbreviations. Required.
    open report - open the Excel report file when finished. True (default) or False. Optional.
    
Enclose any arguments with spaces within double-quotes.
"""

import os, sys
import string
import arcpy
import requests
import json
from distutils.util import strtobool
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill, Alignment

# pyinstaller
# comment out arcpy
# find/replace arcpy.AddMessage with print
# pyinstaller –F GeMS_GeolexCheck_AGP2.py
# use conda environment names-check
#import temptree

versionString = "GeMS_GeolexCheck_AGP2.py, 10/20/2020"

# initialize empty list to collect usage matches in order to avoid 
# displaying redundant matches.
usages = []
     
# STRING AND USAGE
def sanitize_text(usage_text):
    """Clean up usage text so that it only includes alphanumeric characters """
    # entire text might be enclosed in parantheses
    if usage_text[0] == '(' and usage_text[-1] == ')':
        usage_text = usage_text[1:len(usage_text) - 1]
    
    # remove all paranthetical phrases, inside of parantheses or square brackets
    usage_text = re.sub(r'\([^)]*\)', '', usage_text)
    # or square brackets
    usage_text = re.sub(r'\[[^)]*\]', '', usage_text)
        
    # strip all non-alphanumeric characters from string except for periods and apostrophes
    # to allow names like St. John's
    usage_text = re.sub(r"[^\w.']+", ' ', usage_text)
    
    # replace multiple spaces with a single space
    usage_text = re.sub('[\t+\s+]',' ',usage_text)
    
    return usage_text.strip()
    
def sanitize_matches(list1, nametext):
    """Remove names that occur in other, longer names, eg.,
        Saddle in Saddle Mountains or Basin in Basin City, etc.
        and returns a list; [position of Geolex name in fullname, Geolex name]
        position is first item in list so that the list can be sorted"""
        
    # remove duplicates
    list1 = set(list1)
    
    # make a dictionary with keys from the list
    match_d = dict.fromkeys(list1, '')
    list2 = list1
    for item1 in list1:
        for item2 in list2:
            if (item2.startswith(item1 + ' ') or
                item2.endswith(' ' + item1) or
                item2.find(' ' + item1 + ' ') != -1 and
                len(item2) > len(item1)):
                match_d[item1] = 'delete'

    # and now sort by position in the name text
    list3 = [item for item in list1 if not match_d[item] == 'delete']
    list4 = []
    for name in list3:
        list4.append([nametext.find(name), name])
        
    return sorted(list4)
    
def check_usage(glx_usage, fooname):
    """Look for exact or partial match between the supplied name and usage"""
    usage = sanitize_text(glx_usage)
    if usage == fooname.strip() or fooname.strip().find(usage) >= 0:
        return True
    else:
        # there is no usage match
        return False
        
def ext_check(states_list, fn_ext):
    states_list = [foo.lower() for foo in states_list]
    fn_ext = [bar.lower() for bar in fn_ext]
    if set(states_list).intersection(set(fn_ext)):
        return True
    else:
        return False

def parse_age(age_str):
    return age_str.replace('\r\n', '\n')
    
# API 
def units_query(fn):
    """Prepare and send the GET request"""
    units_api = r"https://ngmdb-dev.usgs.gov/db/apiv1/geolex/units/?"
    params = {'units_in': fn}
    response = requests.get(units_api, params)  #.text
    
    return response.json()['results']
    
# EXCEL
def frame_it(d_path, ext_format):
    """convert table to pandas dataframe
       gdbs and excel files need special consideration
       but character-delimited text files can be opened with read_table
    """
    types = {'hierarchykey': str, 'name': str, 'fullname': str}
    
    # attempt to allow all cases of column names
    flds = ['hierarchykey', 'name', 'age', 'fullname', 'mapunit']
    if ext_format == 'gdb':
        # gdb table has to first be converted to csv using gdal command ogr2ogr
        # write it to a temporary directory
        t_dir = tempfile.mkdtemp()
        t_dmu = os.path.join(t_dir, 'dmu.csv')
        gdb_p = os.path.dirname(d_path)
        dmu_table = os.path.basename(d_path)
        ogr_com = f'ogr2ogr -f CSV {t_dmu} {gdb_p} {dmu_table}'
        os.system(ogr_com)
         
        dmu_df = pd.read_csv(t_dmu, usecols=lambda x: x.lower() in flds, dtype=types)
        
        # delete the temp files and directory
        rmtree(t_dir)
    
    elif ext_format == 'xls':
        file = os.path.dirname(d_path)
        sheet = os.path.basename(d_path)
        if sheet[-1:] == '$':
            sheet = sheet[:-1]
        dmu_df = pd.read_excel(file, sheet_name=sheet, engine="openpyxl", usecols=lambda x: x.lower() in flds, dtype=types)
    
    elif ext_format == 'csv':
        dmu_df = pd.read_csv(d_path, usecols=lambda x: x.lower() in flds, dtype=types, keep_default_na=False)
    
    else:
        dmu_df = pd.read_table(d_path, usecols=lambda x: x.lower() in flds, dtype=type, keep_default_na=False)
    
    # smash all column names to lower case because we can't be sure of the case
    # in the input dmu
    dmu_df.columns = [c.lower() for c in dmu_df.columns]
    
    return dmu_df

def link(cell, link, display='link'):
    cell.value = '=HYPERLINK("%s", "%s")' % (link, display)
    cell.font = Font(u='single', color='0000EE')

def format_excel(xlf):
    """Format the output Excel spreadsheet"""
    wb = load_workbook(filename = xlf)
    ws = wb['Sheet1']
    #ws.delete_cols(1)

    # this is the regular Excel border style but it has to be applied after
    # applying the colors, which erases all borders
    border = Border(left=Side(border_style='thin', color='D3D3D3'),
                    right=Side(border_style='thin', color='D3D3D3'),
                    top=Side(border_style='thin', color='D3D3D3'),
                    bottom=Side(border_style='thin', color='D3D3D3'))
    
    # black outline border for header column names
    blackBorder = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    greenFill = PatternFill(start_color='ebf1de',
                            end_color = 'ebf1de',
                            fill_type='solid')

    yellowFill = PatternFill(start_color='ffff99',
                            end_color = 'ffff99',
                            fill_type='solid')

    orangeFill = PatternFill(start_color='fabf8f',
                            end_color = 'fabf8f',
                            fill_type='solid')

    ws.insert_rows(1)
    ws['A1'] = "DMU Contents"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    ws['G1'] = "Geolex Results"
    ws['G1'].font = Font(bold=True)
    ws['G1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('G1:L1')

    ws['M1'] = "Author Review"
    ws['M1'].font = Font(bold=True)
    ws['M1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('M1:P1')

    maxRow = ws.max_row + 1
    
    # color the sections
    for colNum in range(1, 7):
        for rowNum in range(1, maxRow):
            ws.cell(row=rowNum, column=colNum).fill = greenFill
            ws.cell(row=rowNum, column=colNum).border = border
            
    for colNum in range(7, 13):
        for rowNum in range(1, maxRow):
            ws.cell(row=rowNum, column=colNum).fill = yellowFill
            ws.cell(row=rowNum, column=colNum).border = border
            
    for colNum in range(13, 17):
        for rowNum in range(1, maxRow):
            ws.cell(row=rowNum, column=colNum).fill = orangeFill
            ws.cell(row=rowNum, column=colNum).border = border
            
    # apply hyperlink styling to column K, URL
    for rowNum in range(3, maxRow):
        ws_cell = ws.cell(row=rowNum, column=12)
        if not ws_cell.value is None:
            link(ws_cell, ws_cell.value, ws_cell.value)
    
    # re-apply black borders around the header cells
    for colNum in range(1, 17):
        for rowNum in range(1, 3):
            ws.cell(row=rowNum, column=colNum).border = blackBorder
    
    # adjust the width of the cells
    for i in list(string.ascii_uppercase[0:15]):
        ws.column_dimensions[i].width = 12
        
    # materialized paths often get imported to Excel as dates.
    # try to ensure HierarchyKey as text, not date
    for rowNum in range (1, maxRow):
        ws.cell(row=rowNum, column=1).number_format = '@'
        
    wb.save(xlf)

 # START
#------------------------------------------------------------------------
if len(sys.argv) == 1:
    print(__doc__)
    quit()

arcpy.AddMessage(versionString)

# collect the path to the DMU table
dmu = sys.argv[1]

if '.xlsx' in dmu or '.xls' in dmu:
    if not dmu[-1:] == '$':
        dmu = dmu + '$'
        
# get parent directory
dmu_home = os.path.dirname(dmu)
# figure out what the file format is
if os.path.splitext(dmu_home)[1] == '.gdb':
    out_name = os.path.basename(dmu_home)[:-4]
    dmu_home = os.path.dirname(dmu_home)
    dmu_df = frame_it(dmu, 'gdb')

elif os.path.splitext(dmu_home)[1] == '.xlsx':
    out_name = os.path.basename(dmu_home)[:-5]
    dmu_home = os.path.dirname(dmu_home)
    dmu_df = frame_it(dmu, 'xls')

elif os.path.splitext(dmu)[1] == '.csv':
    out_name = os.path.basename(dmu)[:-4]
    dmu_df = frame_it(dmu, 'csv')

elif os.path.splitext(dmu)[1] == '.txt':
    out_name = os.path.basename(dmu)[:-4]
    dmu_df = frame_it(dmu, 'txt')

else:
    arcpy.AddMessage("The DMU file cannot be read\n" +
    "Choose an ESRI file geodatabase table, an Excel spreadsheet,\n" +
    "a comma-delimited text file, or a tab-delimited text file")
    
# collect and clean the extent of the DMU. 
# can be single state or list of states, comma separated,
# can be upper or lower case
dmu_str = sys.argv[2]
dmu_str = dmu_str.strip("\'")
dmu_str = dmu_str.replace(" ", "")
dmu_exts = re.split(';|,',dmu_str)
 
# open the report after running?
if len(sys.argv) == 4:
    open_xl = bool(strtobool(sys.argv[3]))
else:
    open_xl = True

# units table of geolex db
this_py = os.path.realpath(__file__)
geolex_db = os.path.join(os.path.dirname(this_py),'..', 'Resources', 'geolex_units.json')

# set up a pandas data frame
d={}
df = pd.DataFrame(columns=['HierarchyKey', 'MapUnit', 'Name', 'Fullname', 'Age', 'Extent', # DMU Contents
                    'GeolexID', 'Name', 'Usage', 'Age', 'Extent', 'URL',           # Geolex Results
                    'Extent Match?', 'Usage Match?', 'Remarks', 'References'])              # Author Review
                    
df['HierarchyKey'] = df['HierarchyKey'].astype('object')

fields = ['hierarchykey', 'mapunit', 'name', 'fullname', 'age']
#rows = sorted(arcpy.da.SearchCursor(dmu, fields))

n = 0
for row in dmu_df.itertuples():
    # only proceed if there is either a Name or Fullname. This will check for Geolex names in headings
    if row.name or row.fullname: 
        # get some values from the input
        # map unit abbreviation
        mu = row.mapunit 
        #arcpy.AddMessage(f'Checking Name and Fullname for {mu}')
        
        # short map unit name
        if not pd.isna(row.name):
            sn = row.name
            #arcpy.AddMessage(f"Name = {sn}")
            sn_subbed = sanitize_text(sn).strip().lower()
            sn_lower = sn.lower()
        else:
            sn = ''
            sn_subbed = ''
            sn_lower = ''
        
        # full map unit name
        if not pd.isna(row.fullname):
            fn = row.fullname
            #arcpy.AddMessage(f"Fullname = {fn}")
            fn_subbed = sanitize_text(fn).strip().lower()
            fn_lower = fn.lower()
        else:
            fn =''
            fn_subbed = ''
            fn_lower = ''
           
        age = row.age
        
        # pandas might read in materialized paths as dates.
        # try to ensure HierarchyKey is read in as text, not date or a number_format
        hkey = str(row.hierarchykey)
        
        # Collect the geolex names that are in fullname and name
        # Case 1: set of geolex names that are in name same as those in fullname
        #   use the set of geolex names
        # Case 2: there are more geolex names in fullname than in name
        #   use only the geolex names that are in name
        # Case 3: no geolex names in name but there are geolex names in fullname
        #   use the set of geolex names that are in fullname

        # pass the current name and fullname to the api-query function
        sn_matches = None
        fn_matches = None
        
        sn_results = None
        fn_results = None
        if sn:
            sn_results = units_query(sn)
        
        if fn:
            fn_results = units_query(fn)
            
        # if there are name and fullname matches, take the intersection of the sets
        if (sn_results and fn_results) or (sn_results and not fn_results):
            results = sn_results
            #arcpy.AddMessage(results)
            check_name = sn
            
        # if there are only fullname matches, take those
        elif fn_results and not sn_results:
            results = fn_results 
            #arcpy.AddMessage("case 2")
            check_name = fn
       
        # if none of those, there are no matches
        else:
            results = None
            
        # initiate this row filling out the first 6 columns
        # needs to be defined outside of 'if matches' statement below for the case where
        # there are no valid matches
        #unit_list = [mu, fn, fm, age, ext]
        unit_list = [hkey, mu, sn, fn, age, ', '.join(dmu_exts)]
        
        # initialize counter to determine contents of unit_list as matches are recorded
        i = 0 
        
        # initialize incrementing character to append to HierarchyKey in case there are more than one usage
        # this will allow the output table to be sorted on HierarchyKey correctly 
        ch = "a"
        
        if results:
            arcpy.AddMessage(f"Looking for GEOLEX names in {check_name}")
            names_only = [result['unit_name'] for result in results]
            names_only = sanitize_matches(names_only, check_name)
        else:
            names_only = None
        
        if names_only:
            for name in names_only:
                for r in [result for result in results if result['unit_name'] == name[1]]:
                    arcpy.AddMessage(f"Evaluating usages for {name[1]}")
                    glx_id = r['id']
                    glx_name = name[1]
                    glx_age = parse_age(r['age_description'][0])
                    glx_url = r['url']
                   
                    # begin iterating the usages
                    n = 0
                    for usage_i in r['usages']:
                        if usage_i['usage'].lower().find('recognized') == -1 or usage_i['usage'].lower().find('notably') == -1:
                            # check the extent
                            ext_bool = ext_check(usage_i['states'], dmu_exts)
                            if ext_bool:
                                ext_str = 'yes'
                            else:
                                ext_str = 'no'
                                
                            glx_ext = ', '.join(usage_i['states'])

                            # re-write unit_list if the first row has already been written
                            if i == 1:
                                sub_hkey = f"{hkey}-{ch}"
                                # 4 of the first 5 columns are empty in all rows following the first row
                                # append an incrementing character to hkey so that the table can be sorted properly
                                unit_list = [sub_hkey,'','','','','']
                                ch = chr(ord(ch) + 1)
                             
                            if n == 1:
                                glx_id = ''
                                glx_name = ''
                                glx_url = ''
                       
                            # extend the list with Geolex results
                            unit_list.extend([glx_id, glx_name, usage_i['usage'], glx_age, glx_ext, glx_url, ext_str, '', '', ''])
                            
                            # add list to dataframe
                            unit_series = pd.Series(unit_list, index = df.columns)
                            df = df.append(unit_series, ignore_index=True)
                            
                            n = 1
                            i = 1
                    
        # there is no match
        else:
            nomatch = unit_list.extend(['', '', '', '', '', '', 'no', '', '', ''])
            
            # add list to dataframe
            unit_series = pd.Series(unit_list, index = df.columns)
            df = df.append(unit_series, ignore_index=True)
   
xl_path = os.path.join(dmu_home, f'{out_name}_namescheck.xlsx')
arcpy.AddMessage(f"Saving {xl_path}")
if os.path.exists(xl_path):
    os.remove(xl_path)

df.to_excel(xl_path, freeze_panes = (2,0), index=False, engine="openpyxl")
format_excel(xl_path)
if open_xl == True:
    os.startfile(xl_path)